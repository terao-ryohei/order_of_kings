import openpyxl
import os

def esc(s):
    if s is None:
        return None
    return str(s).replace("'", "''")

script_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(script_dir, '..', '武将データ.xlsx')
out_path = os.path.join(script_dir, '..', 'server', 'drizzle', 'seed_warrior_roles.sql')

wb = openpyxl.load_workbook(xlsx_path, read_only=False)

# 武将名→id マップを武将ALLシートから構築
ws_all = wb['武将ALL']
warrior_map = {}  # name -> id
all_rows = list(ws_all.iter_rows(min_row=2, values_only=True))
for i, row in enumerate(all_rows):
    if row[0] is not None and row[1] is not None:  # name と reading が両方あれば有効武将
        name = str(row[0]).strip()
        warrior_id = i + 1
        warrior_map[name] = warrior_id

print(f"武将マップ: {len(warrior_map)}件")

# 役割シート (武将ALL, スキルALL, ゲーム仕様 以外のシート)
exclude_sheets = {'武将ALL', 'スキルALL', 'ゲーム仕様'}
role_sheets = [s for s in wb.sheetnames if s not in exclude_sheets]
print(f"役割シート ({len(role_sheets)}枚): {role_sheets}")

warrior_roles_sql = []
not_found = []

for sheet_name in role_sheets:
    ws = wb[sheet_name]
    role = esc(sheet_name)
    sheet_warriors = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        reading = row[1]
        if name is None:
            continue
        name = str(name).strip()
        # 有効な武将行かチェック: reading が存在すること
        if reading is None:
            continue
        if name in warrior_map:
            warrior_id = warrior_map[name]
            warrior_roles_sql.append(
                f"INSERT OR IGNORE INTO warrior_roles (warrior_id, role) "
                f"VALUES ({warrior_id}, '{role}');"
            )
            sheet_warriors += 1
        else:
            not_found.append(f"{sheet_name}/{name}")
    print(f"  {sheet_name}: {sheet_warriors}件")

wb.close()

with open(out_path, 'w', encoding='utf-8') as f:
    f.write('-- 役割別シートから warrior_roles を生成\n')
    f.write('-- GENERATED_AT: 2026-03-26\n\n')
    for sql in warrior_roles_sql:
        f.write(sql + '\n')

print(f"\nGenerated {len(warrior_roles_sql)} warrior_roles")
if not_found:
    print(f"武将マップ未マッチ ({len(not_found)}件): {not_found[:10]}")
